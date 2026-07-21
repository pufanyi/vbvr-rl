#!/usr/bin/env python3
"""Collect regressed VBVR tasks and their paired videos into a review directory."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path
from statistics import fmean
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TASK_WORKBOOK_RELATIVE_PATH = Path("scores/eval_1024x1024_161f_5s_task_scores.xlsx")
RESULT_JSON_RELATIVE_PATH = Path("scores/eval_1024x1024_161f_5s_vbvr_results.json")
EVAL_SAMPLES_RELATIVE_PATH = Path("eval_samples.json")
SCORED_VIDEO_DIR = "eval_1024x1024_161f_5s"
RAW_VIDEO_DIR = "generated_256x256x161"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--baseline-root", type=Path, required=True)
    parser.add_argument("--candidate-root", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--baseline-label", default="baseline")
    parser.add_argument("--candidate-label", default="candidate")
    parser.add_argument(
        "--materialize",
        choices=("copy", "hardlink", "symlink"),
        default="copy",
        help="How videos are materialized in the review directory (default: copy).",
    )
    return parser.parse_args()


def load_json(path: Path) -> Any:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_task_rows(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Task Scores"]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    expected = ("Split", "Category", "Task Name", "Sample Count", "Average Score")
    if header != expected:
        raise ValueError(f"Unexpected task workbook header in {path}: {header!r}")

    tasks: dict[str, dict[str, Any]] = {}
    for split, category, task_name, sample_count, average_score in rows:
        if task_name in tasks:
            raise ValueError(f"Duplicate task {task_name!r} in {path}")
        tasks[str(task_name)] = {
            "split": str(split),
            "category": str(category),
            "sample_count": int(sample_count),
            "average_score": float(average_score),
        }
    return tasks


def load_result_samples(path: Path) -> tuple[dict[str, Any], dict[tuple[str, str], dict[str, Any]]]:
    result = load_json(path)
    indexed: dict[tuple[str, str], dict[str, Any]] = {}
    for sample in result["samples"]:
        video_id = Path(sample["video_file"]).stem
        key = (sample["task_name"], video_id)
        if key in indexed:
            raise ValueError(f"Duplicate scored sample {key!r} in {path}")
        if sample["error"] is not None:
            raise ValueError(f"Scorer error for {key!r} in {path}: {sample['error']}")
        indexed[key] = sample
    return result, indexed


def load_eval_samples(path: Path) -> dict[tuple[str, str], dict[str, Any]]:
    indexed: dict[tuple[str, str], dict[str, Any]] = {}
    for sample in load_json(path):
        key = (sample["task_name"], str(sample["video_idx"]))
        if key in indexed:
            raise ValueError(f"Duplicate eval sample {key!r} in {path}")
        indexed[key] = sample
    return indexed


def validate_inputs(
    baseline_root: Path,
    candidate_root: Path,
    baseline_tasks: dict[str, dict[str, Any]],
    candidate_tasks: dict[str, dict[str, Any]],
    baseline_scores: dict[tuple[str, str], dict[str, Any]],
    candidate_scores: dict[tuple[str, str], dict[str, Any]],
    baseline_eval_samples: dict[tuple[str, str], dict[str, Any]],
    candidate_eval_samples: dict[tuple[str, str], dict[str, Any]],
) -> None:
    if baseline_tasks.keys() != candidate_tasks.keys():
        missing_candidate = sorted(baseline_tasks.keys() - candidate_tasks.keys())
        missing_baseline = sorted(candidate_tasks.keys() - baseline_tasks.keys())
        raise ValueError(
            f"Task sets differ; missing from candidate={missing_candidate}, missing from baseline={missing_baseline}"
        )
    if baseline_scores.keys() != candidate_scores.keys():
        raise ValueError("The two result JSON files do not contain the same scored sample keys")
    if baseline_scores.keys() != baseline_eval_samples.keys():
        raise ValueError("Baseline result JSON and eval_samples.json contain different sample keys")
    if candidate_scores.keys() != candidate_eval_samples.keys():
        raise ValueError("Candidate result JSON and eval_samples.json contain different sample keys")

    for key in baseline_eval_samples:
        baseline = baseline_eval_samples[key]
        candidate = candidate_eval_samples[key]
        for field in ("name", "task_name", "video_idx", "domain", "prompt", "image"):
            if baseline[field] != candidate[field]:
                raise ValueError(f"Eval sample field {field!r} differs for {key!r}")

    for task_name, baseline_task in baseline_tasks.items():
        candidate_task = candidate_tasks[task_name]
        for field in ("split", "category", "sample_count"):
            if baseline_task[field] != candidate_task[field]:
                raise ValueError(f"Task metadata field {field!r} differs for {task_name!r}")

        baseline_samples = [
            sample["score"] for (sample_task, _), sample in baseline_scores.items() if sample_task == task_name
        ]
        candidate_samples = [
            sample["score"] for (sample_task, _), sample in candidate_scores.items() if sample_task == task_name
        ]
        if len(baseline_samples) != baseline_task["sample_count"]:
            raise ValueError(f"Baseline sample count mismatch for {task_name!r}")
        if len(candidate_samples) != candidate_task["sample_count"]:
            raise ValueError(f"Candidate sample count mismatch for {task_name!r}")
        if not math.isclose(
            fmean(baseline_samples),
            baseline_task["average_score"],
            rel_tol=0.0,
            abs_tol=1e-12,
        ):
            raise ValueError(f"Baseline workbook average does not match result JSON for {task_name!r}")
        if not math.isclose(
            fmean(candidate_samples),
            candidate_task["average_score"],
            rel_tol=0.0,
            abs_tol=1e-12,
        ):
            raise ValueError(f"Candidate workbook average does not match result JSON for {task_name!r}")

    for root, scores in ((baseline_root, baseline_scores), (candidate_root, candidate_scores)):
        for (task_name, video_id), sample in scores.items():
            video_file = f"{video_id}.mp4"
            scored_path = root / SCORED_VIDEO_DIR / sample["folder"] / task_name / video_file
            raw_path = root / RAW_VIDEO_DIR / sample["folder"] / task_name / video_file
            if not scored_path.is_file():
                raise FileNotFoundError(scored_path)
            if not raw_path.is_file():
                raise FileNotFoundError(raw_path)
            recorded_path = Path(sample["video_path"])
            if recorded_path.resolve() != scored_path.resolve():
                raise ValueError(
                    f"Recorded scorer path differs from expected prepared video for {(task_name, video_id)!r}: "
                    f"{recorded_path} != {scored_path}"
                )


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def materialize_file(source: Path, destination: Path, mode: str) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if mode == "copy":
        shutil.copy2(source, destination)
    elif mode == "hardlink":
        os.link(source, destination)
    else:
        destination.symlink_to(os.path.relpath(source, start=destination.parent))


def score_filename(video_id: str, score: float) -> str:
    return f"{video_id}__score_{score:.6f}.mp4"


def write_task_readme(
    path: Path,
    task: dict[str, Any],
    samples: list[dict[str, Any]],
    baseline_label: str,
    candidate_label: str,
) -> None:
    lines = [
        f"# {task['task_name']}",
        "",
        f"- Split: `{task['split']}`",
        f"- Category: `{task['category']}`",
        f"- {baseline_label} task average: `{task['baseline_average']:.12f}`",
        f"- {candidate_label} task average: `{task['candidate_average']:.12f}`",
        f"- Delta (candidate - baseline): `{task['delta']:+.12f}`",
        "",
        "| Video | Baseline score | Candidate score | Delta |",
        "|---|---:|---:|---:|",
    ]
    for sample in samples:
        lines.append(
            f"| {sample['video_id']} | {sample['baseline_score']:.12f} | "
            f"{sample['candidate_score']:.12f} | {sample['delta']:+.12f} |"
        )
    lines.extend(
        [
            "",
            "The `scored_1024x1024_161f_5s/` videos are the exact prepared videos scored by EvalKit. "
            "The `raw_generated_256x256x161/` videos are the corresponding raw model outputs. "
            "Filenames show scores rounded to six decimal places; `scores.csv` preserves full precision.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def autosize_sheet(sheet: Any, widths: dict[str, float] | None = None) -> None:
    widths = widths or {}
    for index, column in enumerate(sheet.iter_cols(), start=1):
        letter = get_column_letter(index)
        if letter in widths:
            sheet.column_dimensions[letter].width = widths[letter]
            continue
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 60)


def style_sheet(sheet: Any, delta_column: str | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if delta_column is not None and sheet.max_row >= 2:
        sheet.conditional_formatting.add(
            f"{delta_column}2:{delta_column}{sheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="num",
                mid_value=0,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )


def write_workbook(
    path: Path,
    dropped_rows: list[dict[str, Any]],
    per_video_rows: list[dict[str, Any]],
    all_task_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    dropped_sheet = workbook.active
    dropped_sheet.title = "Dropped Tasks"
    dropped_headers = [
        "Drop Rank",
        "Split",
        "Category",
        "Task Name",
        "Sample Count",
        "Baseline Average",
        "Candidate Average",
        "Delta",
        "Relative Change (%)",
        "Task Folder",
    ]
    dropped_sheet.append(dropped_headers)
    for row in dropped_rows:
        dropped_sheet.append(
            [
                row["drop_rank"],
                row["split"],
                row["category"],
                row["task_name"],
                row["sample_count"],
                row["baseline_average"],
                row["candidate_average"],
                row["delta"],
                row["relative_change_percent"],
                row["task_folder"],
            ]
        )
        folder_cell = dropped_sheet.cell(dropped_sheet.max_row, 10)
        folder_cell.hyperlink = f"{row['task_folder']}/scores.csv"
        folder_cell.style = "Hyperlink"
    style_sheet(dropped_sheet, "H")
    autosize_sheet(dropped_sheet, {"D": 78, "J": 90})
    for row in range(2, dropped_sheet.max_row + 1):
        for column in ("F", "G", "H"):
            dropped_sheet[f"{column}{row}"].number_format = "0.000000000000"
        dropped_sheet[f"I{row}"].number_format = "0.000000"

    video_sheet = workbook.create_sheet("Per-Video Scores")
    video_headers = [
        "Drop Rank",
        "Split",
        "Category",
        "Task Name",
        "Video ID",
        "Baseline Score",
        "Candidate Score",
        "Delta",
        "Baseline Scored Video",
        "Candidate Scored Video",
        "Baseline Raw Video",
        "Candidate Raw Video",
        "Input Image",
        "Prompt",
    ]
    video_sheet.append(video_headers)
    for row in per_video_rows:
        video_sheet.append(
            [
                row["drop_rank"],
                row["split"],
                row["category"],
                row["task_name"],
                row["video_id"],
                row["baseline_score"],
                row["candidate_score"],
                row["delta"],
                row["baseline_scored_video"],
                row["candidate_scored_video"],
                row["baseline_raw_video"],
                row["candidate_raw_video"],
                row["input_image"],
                row["prompt"],
            ]
        )
        for column in range(9, 13):
            cell = video_sheet.cell(video_sheet.max_row, column)
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
    style_sheet(video_sheet, "H")
    autosize_sheet(video_sheet, {"D": 78, "H": 18, "I": 80, "J": 80, "K": 80, "L": 80, "M": 80, "N": 100})
    for row in range(2, video_sheet.max_row + 1):
        for column in ("F", "G", "H"):
            video_sheet[f"{column}{row}"].number_format = "0.000000000000"

    all_sheet = workbook.create_sheet("All Tasks")
    all_headers = [
        "Split",
        "Category",
        "Task Name",
        "Sample Count",
        "Baseline Average",
        "Candidate Average",
        "Delta",
        "Status",
    ]
    all_sheet.append(all_headers)
    for row in all_task_rows:
        all_sheet.append(
            [
                row["split"],
                row["category"],
                row["task_name"],
                row["sample_count"],
                row["baseline_average"],
                row["candidate_average"],
                row["delta"],
                row["status"],
            ]
        )
    style_sheet(all_sheet, "G")
    autosize_sheet(all_sheet, {"C": 78})
    for row in range(2, all_sheet.max_row + 1):
        for column in ("E", "F", "G"):
            all_sheet[f"{column}{row}"].number_format = "0.000000000000"

    workbook.save(path)


def main() -> None:
    args = parse_args()
    baseline_root = args.baseline_root.resolve()
    candidate_root = args.candidate_root.resolve()
    output_dir = args.output_dir.resolve()

    required_paths = [
        baseline_root / TASK_WORKBOOK_RELATIVE_PATH,
        candidate_root / TASK_WORKBOOK_RELATIVE_PATH,
        baseline_root / RESULT_JSON_RELATIVE_PATH,
        candidate_root / RESULT_JSON_RELATIVE_PATH,
        baseline_root / EVAL_SAMPLES_RELATIVE_PATH,
        candidate_root / EVAL_SAMPLES_RELATIVE_PATH,
    ]
    missing = [path for path in required_paths if not path.is_file()]
    if missing:
        raise FileNotFoundError(f"Missing required inputs: {missing}")
    if output_dir.exists():
        raise FileExistsError(f"Refusing to overwrite existing output directory: {output_dir}")

    baseline_tasks = load_task_rows(baseline_root / TASK_WORKBOOK_RELATIVE_PATH)
    candidate_tasks = load_task_rows(candidate_root / TASK_WORKBOOK_RELATIVE_PATH)
    baseline_result, baseline_scores = load_result_samples(baseline_root / RESULT_JSON_RELATIVE_PATH)
    candidate_result, candidate_scores = load_result_samples(candidate_root / RESULT_JSON_RELATIVE_PATH)
    baseline_eval_samples = load_eval_samples(baseline_root / EVAL_SAMPLES_RELATIVE_PATH)
    candidate_eval_samples = load_eval_samples(candidate_root / EVAL_SAMPLES_RELATIVE_PATH)
    validate_inputs(
        baseline_root,
        candidate_root,
        baseline_tasks,
        candidate_tasks,
        baseline_scores,
        candidate_scores,
        baseline_eval_samples,
        candidate_eval_samples,
    )

    all_task_rows: list[dict[str, Any]] = []
    for task_name, baseline in baseline_tasks.items():
        candidate = candidate_tasks[task_name]
        delta = candidate["average_score"] - baseline["average_score"]
        status = "dropped" if delta < 0 else "increased" if delta > 0 else "equal"
        all_task_rows.append(
            {
                "split": baseline["split"],
                "category": baseline["category"],
                "task_name": task_name,
                "sample_count": baseline["sample_count"],
                "baseline_average": baseline["average_score"],
                "candidate_average": candidate["average_score"],
                "delta": delta,
                "status": status,
            }
        )
    all_task_rows.sort(key=lambda row: (row["delta"], row["task_name"]))
    dropped_rows = [dict(row) for row in all_task_rows if row["status"] == "dropped"]
    for rank, row in enumerate(dropped_rows, start=1):
        row["drop_rank"] = rank
        row["relative_change_percent"] = (
            row["delta"] / row["baseline_average"] * 100 if row["baseline_average"] != 0 else None
        )
        row["task_folder"] = f"tasks/{rank:03d}__{row['task_name']}"

    output_dir.parent.mkdir(parents=True, exist_ok=True)
    stage_dir = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}.tmp-", dir=output_dir.parent))
    try:
        source_dir = stage_dir / "source_score_files"
        source_dir.mkdir()
        source_copies = {
            "baseline_task_scores.xlsx": baseline_root / TASK_WORKBOOK_RELATIVE_PATH,
            "candidate_task_scores.xlsx": candidate_root / TASK_WORKBOOK_RELATIVE_PATH,
            "baseline_vbvr_results.json": baseline_root / RESULT_JSON_RELATIVE_PATH,
            "candidate_vbvr_results.json": candidate_root / RESULT_JSON_RELATIVE_PATH,
            "baseline_eval_samples.json": baseline_root / EVAL_SAMPLES_RELATIVE_PATH,
            "candidate_eval_samples.json": candidate_root / EVAL_SAMPLES_RELATIVE_PATH,
        }
        for destination_name, source_path in source_copies.items():
            shutil.copy2(source_path, source_dir / destination_name)

        per_video_rows: list[dict[str, Any]] = []
        per_task_samples: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for task in dropped_rows:
            task_name = task["task_name"]
            task_dir = stage_dir / task["task_folder"]
            video_ids = sorted(video_id for sample_task, video_id in baseline_scores if sample_task == task_name)
            for video_id in video_ids:
                key = (task_name, video_id)
                baseline_score = baseline_scores[key]
                candidate_score = candidate_scores[key]
                eval_sample = baseline_eval_samples[key]
                video_file = f"{video_id}.mp4"

                baseline_scored_source = (
                    baseline_root / SCORED_VIDEO_DIR / baseline_score["folder"] / task_name / video_file
                )
                candidate_scored_source = (
                    candidate_root / SCORED_VIDEO_DIR / candidate_score["folder"] / task_name / video_file
                )
                baseline_raw_source = baseline_root / RAW_VIDEO_DIR / baseline_score["folder"] / task_name / video_file
                candidate_raw_source = (
                    candidate_root / RAW_VIDEO_DIR / candidate_score["folder"] / task_name / video_file
                )

                baseline_filename = score_filename(video_id, baseline_score["score"])
                candidate_filename = score_filename(video_id, candidate_score["score"])
                baseline_scored_destination = (
                    task_dir / args.baseline_label / "scored_1024x1024_161f_5s" / baseline_filename
                )
                candidate_scored_destination = (
                    task_dir / args.candidate_label / "scored_1024x1024_161f_5s" / candidate_filename
                )
                baseline_raw_destination = (
                    task_dir / args.baseline_label / "raw_generated_256x256x161" / baseline_filename
                )
                candidate_raw_destination = (
                    task_dir / args.candidate_label / "raw_generated_256x256x161" / candidate_filename
                )
                for source, destination in (
                    (baseline_scored_source, baseline_scored_destination),
                    (candidate_scored_source, candidate_scored_destination),
                    (baseline_raw_source, baseline_raw_destination),
                    (candidate_raw_source, candidate_raw_destination),
                ):
                    materialize_file(source, destination, args.materialize)

                row = {
                    "drop_rank": task["drop_rank"],
                    "split": task["split"],
                    "category": task["category"],
                    "task_name": task_name,
                    "video_id": video_id,
                    "baseline_score": baseline_score["score"],
                    "candidate_score": candidate_score["score"],
                    "delta": candidate_score["score"] - baseline_score["score"],
                    "baseline_dimensions": json.dumps(
                        baseline_score["dimensions"], ensure_ascii=False, sort_keys=True
                    ),
                    "candidate_dimensions": json.dumps(
                        candidate_score["dimensions"], ensure_ascii=False, sort_keys=True
                    ),
                    "baseline_scored_video": baseline_scored_destination.relative_to(stage_dir).as_posix(),
                    "candidate_scored_video": candidate_scored_destination.relative_to(stage_dir).as_posix(),
                    "baseline_raw_video": baseline_raw_destination.relative_to(stage_dir).as_posix(),
                    "candidate_raw_video": candidate_raw_destination.relative_to(stage_dir).as_posix(),
                    "input_image": eval_sample["image"],
                    "prompt": eval_sample["prompt"],
                }
                per_video_rows.append(row)
                per_task_samples[task_name].append(row)

            task_samples = per_task_samples[task_name]
            write_csv(
                task_dir / "scores.csv",
                [
                    "video_id",
                    "baseline_score",
                    "candidate_score",
                    "delta",
                    "baseline_dimensions",
                    "candidate_dimensions",
                    "baseline_scored_video",
                    "candidate_scored_video",
                    "baseline_raw_video",
                    "candidate_raw_video",
                    "input_image",
                    "prompt",
                ],
                [
                    {
                        key: value
                        for key, value in sample.items()
                        if key not in {"drop_rank", "split", "category", "task_name"}
                    }
                    for sample in task_samples
                ],
            )
            task_manifest = {
                **task,
                "baseline_label": args.baseline_label,
                "candidate_label": args.candidate_label,
                "samples": task_samples,
            }
            (task_dir / "task_summary.json").write_text(
                json.dumps(task_manifest, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            write_task_readme(
                task_dir / "README.md",
                task,
                task_samples,
                args.baseline_label,
                args.candidate_label,
            )

        dropped_csv_headers = [
            "drop_rank",
            "split",
            "category",
            "task_name",
            "sample_count",
            "baseline_average",
            "candidate_average",
            "delta",
            "relative_change_percent",
            "task_folder",
        ]
        write_csv(
            stage_dir / "dropped_tasks.csv",
            dropped_csv_headers,
            [{key: row[key] for key in dropped_csv_headers} for row in dropped_rows],
        )
        per_video_csv_headers = [
            "drop_rank",
            "split",
            "category",
            "task_name",
            "video_id",
            "baseline_score",
            "candidate_score",
            "delta",
            "baseline_dimensions",
            "candidate_dimensions",
            "baseline_scored_video",
            "candidate_scored_video",
            "baseline_raw_video",
            "candidate_raw_video",
            "input_image",
            "prompt",
        ]
        write_csv(stage_dir / "per_video_scores.csv", per_video_csv_headers, per_video_rows)
        all_task_csv_headers = [
            "split",
            "category",
            "task_name",
            "sample_count",
            "baseline_average",
            "candidate_average",
            "delta",
            "status",
        ]
        write_csv(stage_dir / "all_tasks.csv", all_task_csv_headers, all_task_rows)
        write_workbook(stage_dir / "comparison.xlsx", dropped_rows, per_video_rows, all_task_rows)

        status_counts = {
            status: sum(row["status"] == status for row in all_task_rows)
            for status in ("dropped", "increased", "equal")
        }
        source_hashes = {name: sha256(path) for name, path in source_copies.items()}
        baseline_overall = float(baseline_result["summary"]["overall"]["mean_score"])
        candidate_overall = float(candidate_result["summary"]["overall"]["mean_score"])
        manifest = {
            "baseline": {
                "label": args.baseline_label,
                "root": str(baseline_root),
                "overall_score": baseline_overall,
            },
            "candidate": {
                "label": args.candidate_label,
                "root": str(candidate_root),
                "overall_score": candidate_overall,
            },
            "criterion": "candidate task average < baseline task average (strict floating-point comparison)",
            "materialize_mode": args.materialize,
            "task_count": len(all_task_rows),
            "status_counts": status_counts,
            "dropped_sample_count": len(per_video_rows),
            "source_sha256": source_hashes,
        }
        (stage_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        readme_lines = [
            "# VBVR-Pro task regression review",
            "",
            f"- Baseline: `{baseline_root}`",
            f"- Candidate: `{candidate_root}`",
            f"- Overall score: `{baseline_overall:.12f}` → `{candidate_overall:.12f}` "
            f"(`{candidate_overall - baseline_overall:+.12f}`)",
            f"- Compared tasks: `{len(all_task_rows)}`",
            f"- Dropped / increased / equal: `{status_counts['dropped']} / "
            f"{status_counts['increased']} / {status_counts['equal']}`",
            "- Drop criterion: candidate task average is strictly less than baseline task average; "
            "no epsilon threshold was applied.",
            "",
            "## Contents",
            "",
            "- `comparison.xlsx`: dropped-task summary, all paired per-video scores, and all-task deltas.",
            "- `dropped_tasks.csv`: the dropped tasks, ordered from the largest drop to the smallest.",
            "- `per_video_scores.csv`: all paired samples for dropped tasks, including full-precision "
            "scores and prompts.",
            "- `all_tasks.csv`: all 100 task deltas and their dropped/increased/equal status.",
            "- `tasks/NNN__<task_name>/`: one review directory per dropped task.",
            "- `source_score_files/`: copies of the two source workbooks, result JSON files, "
            "and eval sample manifests.",
            "",
            "Each task directory contains both runs' raw 256×256×161 generated videos and the exact "
            "1024×1024×161/≤5s prepared videos scored by EvalKit. Video filenames show the corresponding "
            "score rounded to six decimal places; CSV/XLSX files retain full precision.",
            "",
        ]
        (stage_dir / "README.md").write_text("\n".join(readme_lines), encoding="utf-8")

        expected_pairs = sum(row["sample_count"] for row in dropped_rows)
        if len(per_video_rows) != expected_pairs:
            raise RuntimeError(f"Expected {expected_pairs} paired samples, wrote {len(per_video_rows)}")
        expected_videos = expected_pairs * 4
        actual_videos = sum(1 for path in (stage_dir / "tasks").rglob("*.mp4"))
        if actual_videos != expected_videos:
            raise RuntimeError(f"Expected {expected_videos} materialized videos, found {actual_videos}")
        if output_dir.exists():
            raise FileExistsError(f"Output appeared while staging; refusing to overwrite: {output_dir}")
        stage_dir.rename(output_dir)
    except BaseException:
        shutil.rmtree(stage_dir, ignore_errors=True)
        raise

    print(
        json.dumps(
            {
                "output_dir": str(output_dir),
                "dropped_tasks": len(dropped_rows),
                "paired_samples": len(per_video_rows),
                "materialized_videos": len(per_video_rows) * 4,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
