from __future__ import annotations

import argparse
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.eval.evaluation_provenance import verify_recorded_manifest

SCORE_PROVENANCE_NAME = "score-provenance.json"


@dataclass(frozen=True)
class RunResult:
    name: str
    result_path: Path
    summary: dict[str, Any]
    task_scores: dict[str, float]
    task_meta: dict[str, tuple[str, str]]
    task_counts: dict[str, int]
    evalkit_revision: str
    evalkit_source_sha256: str

    @property
    def overall(self) -> float:
        return float(self.summary["overall"]["mean_score"])

    @property
    def in_domain(self) -> float:
        return float(self.summary["In_Domain"]["mean_score"])

    @property
    def out_of_domain(self) -> float:
        return float(self.summary["Out_of_Domain"]["mean_score"])

    @property
    def sample_count(self) -> int:
        return int(self.summary["overall"]["num_samples"])


def _load_run(
    run_dir: Path,
    expected_samples: int,
    expected_evalkit_source_sha256: str | None = None,
) -> RunResult:
    provenance_path = run_dir / SCORE_PROVENANCE_NAME
    if not provenance_path.is_file():
        raise ValueError(f"missing {SCORE_PROVENANCE_NAME}")
    provenance = json.loads(provenance_path.read_text(encoding="utf-8"))
    if provenance.get("schema_version") != 1:
        raise ValueError("unsupported or missing score provenance schema")
    if provenance.get("stage") != "vbvr-pro-score":
        raise ValueError("score provenance stage is not vbvr-pro-score")

    values = provenance.get("values", {})
    if values.get("state") != "complete":
        raise ValueError("score provenance is not complete")
    evalkit_revision = values.get("evalkit_revision")
    evalkit_source_sha256 = values.get("evalkit_source_sha256")
    if not isinstance(evalkit_revision, str) or not evalkit_revision:
        raise ValueError("score provenance is missing evalkit_revision")
    if not isinstance(evalkit_source_sha256, str) or len(evalkit_source_sha256) != 64:
        raise ValueError("score provenance is missing evalkit_source_sha256")
    if expected_evalkit_source_sha256 and evalkit_source_sha256 != expected_evalkit_source_sha256:
        raise ValueError(
            f"scorer fingerprint {evalkit_source_sha256} does not match expected {expected_evalkit_source_sha256}"
        )

    recorded_result = provenance.get("output_files", {}).get("result", {})
    recorded_path = recorded_result.get("path")
    if not isinstance(recorded_path, str) or not recorded_path:
        raise ValueError("score provenance does not record its result JSON")
    result_path = Path(recorded_path)
    if result_path.parent.resolve() != (run_dir / "scores").resolve():
        raise ValueError("recorded result JSON is outside this run's scores directory")
    if not result_path.name.endswith("_vbvr_results.json"):
        raise ValueError("recorded result does not use the VBVR result filename contract")

    matches, detail = verify_recorded_manifest(
        provenance_path,
        expected_stage="vbvr-pro-score",
        require_complete=True,
        sections=("output_files",),
    )
    if not matches:
        raise ValueError(detail)

    result = json.loads(result_path.read_text(encoding="utf-8"))
    samples = result.get("samples")
    if not isinstance(samples, list) or len(samples) != expected_samples:
        raise ValueError(f"expected {expected_samples} samples, found {len(samples or [])}")
    errors = [sample for sample in samples if sample.get("error")]
    if errors:
        raise ValueError(f"scorer errors={len(errors)}")

    task_meta: dict[str, tuple[str, str]] = {}
    task_counts: Counter[str] = Counter()
    for index, sample in enumerate(samples):
        try:
            task_name = str(sample["task_name"])
            metadata = (str(sample["split"]), str(sample["category"]))
            float(sample["score"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"sample {index} has invalid task metadata or score") from exc
        previous = task_meta.setdefault(task_name, metadata)
        if previous != metadata:
            raise ValueError(f"inconsistent metadata for task {task_name}")
        task_counts[task_name] += 1

    try:
        summary = result["summary"]
        for key in ("overall", "In_Domain", "Out_of_Domain"):
            float(summary[key]["mean_score"])
            int(summary[key]["num_samples"])
        task_scores = {name: float(score) for name, score in summary["overall"]["by_task"].items()}
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError("result JSON contains an invalid summary") from exc
    if task_scores.keys() != task_meta.keys():
        raise ValueError("task summary does not match the sample task set")

    return RunResult(
        name=run_dir.name,
        result_path=result_path,
        summary=summary,
        task_scores=task_scores,
        task_meta=task_meta,
        task_counts=dict(task_counts),
        evalkit_revision=evalkit_revision,
        evalkit_source_sha256=evalkit_source_sha256,
    )


def _looks_like_run(path: Path) -> bool:
    return any(
        (path / marker).exists()
        for marker in (
            SCORE_PROVENANCE_NAME,
            "generation-provenance.json",
            "preparation-provenance.json",
            "scores",
        )
    )


def discover_runs(
    root: Path,
    expected_samples: int = 500,
    expected_evalkit_source_sha256: str | None = None,
) -> tuple[list[RunResult], list[tuple[str, str]]]:
    if not root.is_dir():
        raise ValueError(f"result root does not exist or is not a directory: {root}")
    candidates = (
        [root]
        if (root / SCORE_PROVENANCE_NAME).is_file()
        else sorted(
            (path for path in root.iterdir() if path.is_dir() and _looks_like_run(path)),
            key=lambda path: path.name,
        )
    )

    runs: list[RunResult] = []
    skipped: list[tuple[str, str]] = []
    for run_dir in candidates:
        try:
            runs.append(_load_run(run_dir, expected_samples, expected_evalkit_source_sha256))
        except (OSError, KeyError, TypeError, ValueError, json.JSONDecodeError) as exc:
            skipped.append((run_dir.name, str(exc)))

    scorer_hashes = {run.evalkit_source_sha256 for run in runs}
    if len(scorer_hashes) > 1:
        raise ValueError(
            "Refusing to mix VBVR-Pro results from multiple scorer fingerprints: " + ", ".join(sorted(scorer_hashes))
        )
    return runs, skipped


def _format_sheet(sheet: Any, widths: list[int], score_columns: tuple[int, ...] = ()) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for column in score_columns:
        for (cell,) in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
            cell.number_format = "0.000000"


def write_workbook(runs: list[RunResult], skipped: list[tuple[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    run_sheet = workbook.active
    run_sheet.title = "Runs"
    run_sheet.append(
        [
            "Run",
            "Overall",
            "In-Domain",
            "Out-of-Domain",
            "Samples",
            "Tasks",
            "EvalKit Revision",
            "EvalKit Source SHA-256",
            "Result JSON",
        ]
    )
    for run in runs:
        run_sheet.append(
            [
                run.name,
                run.overall,
                run.in_domain,
                run.out_of_domain,
                run.sample_count,
                len(run.task_scores),
                run.evalkit_revision,
                run.evalkit_source_sha256,
                str(run.result_path),
            ]
        )
    _format_sheet(run_sheet, [44, 14, 14, 16, 12, 10, 42, 68, 90], (2, 3, 4))

    task_sheet = workbook.create_sheet("Task Scores")
    task_sheet.append(["Run", "Split", "Category", "Task Name", "Samples", "Average Score"])
    split_order = {"In_Domain": 0, "Out_of_Domain": 1}
    for run in runs:
        for task_name in sorted(
            run.task_scores,
            key=lambda task: (split_order.get(run.task_meta[task][0], 2), task),
        ):
            split, category = run.task_meta[task_name]
            task_sheet.append(
                [
                    run.name,
                    split.replace("_", "-"),
                    category,
                    task_name,
                    run.task_counts[task_name],
                    run.task_scores[task_name],
                ]
            )
    _format_sheet(task_sheet, [44, 18, 20, 78, 12, 16], (6,))

    skipped_sheet = workbook.create_sheet("Skipped")
    skipped_sheet.append(["Run", "Reason"])
    for name, reason in skipped:
        skipped_sheet.append([name, reason])
    _format_sheet(skipped_sheet, [44, 90])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_json_summary(runs: list[RunResult], skipped: list[tuple[str, str]], output_path: Path) -> None:
    payload = {
        "schema_version": 1,
        "run_count": len(runs),
        "skipped_count": len(skipped),
        "evalkit_source_sha256": runs[0].evalkit_source_sha256 if runs else None,
        "runs": [
            {
                "name": run.name,
                "overall": run.overall,
                "in_domain": run.in_domain,
                "out_of_domain": run.out_of_domain,
                "sample_count": run.sample_count,
                "task_count": len(run.task_scores),
                "evalkit_revision": run.evalkit_revision,
                "evalkit_source_sha256": run.evalkit_source_sha256,
                "result_json": str(run.result_path),
            }
            for run in runs
        ],
        "skipped": [{"name": name, "reason": reason} for name, reason in skipped],
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def write_text_summary(runs: list[RunResult], skipped: list[tuple[str, str]], output_path: Path) -> None:
    lines = ["Run\tOverall\tIn-Domain\tOut-of-Domain\tSamples"]
    lines.extend(
        f"{run.name}\t{run.overall:.6f}\t{run.in_domain:.6f}\t{run.out_of_domain:.6f}\t{run.sample_count}"
        for run in runs
    )
    if skipped:
        lines.extend(["", f"Skipped/incomplete: {len(skipped)}"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generate_reports(
    root: Path,
    output_dir: Path,
    expected_samples: int = 500,
    expected_evalkit_source_sha256: str | None = None,
) -> tuple[Path, Path, Path, int, int]:
    runs, skipped = discover_runs(
        root,
        expected_samples=expected_samples,
        expected_evalkit_source_sha256=expected_evalkit_source_sha256,
    )
    if not runs:
        raise ValueError(f"No complete provenance-bound results found under {root}")

    workbook_path = output_dir / "vbvr_pro_summary.xlsx"
    json_path = output_dir / "vbvr_pro_summary.json"
    text_path = output_dir / "final_scores.txt"
    write_workbook(runs, skipped, workbook_path)
    write_json_summary(runs, skipped, json_path)
    write_text_summary(runs, skipped, text_path)
    return workbook_path, json_path, text_path, len(runs), len(skipped)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Summarize complete, provenance-bound VBVR-Pro rule-evaluation cells.")
    parser.add_argument("--root", type=Path, required=True, help="One run directory or a parent of sampler cells")
    parser.add_argument("--output-dir", type=Path, default=None, help="Default: <root>/reports")
    parser.add_argument("--expected-samples", type=int, default=500)
    parser.add_argument(
        "--expected-evalkit-source-sha256",
        default=None,
        help="Optionally require this exact scorer-contract fingerprint",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir or args.root / "reports"
    paths = generate_reports(
        args.root,
        output_dir,
        expected_samples=args.expected_samples,
        expected_evalkit_source_sha256=args.expected_evalkit_source_sha256,
    )
    for path in paths[:3]:
        print(path)
    print(f"Complete runs: {paths[3]}; skipped/incomplete runs: {paths[4]}")


if __name__ == "__main__":
    main()
