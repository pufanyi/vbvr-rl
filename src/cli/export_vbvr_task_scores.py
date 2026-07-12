from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export per-task VBVR scores to an Excel workbook.")
    parser.add_argument("result_json", type=Path, help="VBVR result JSON produced by the parallel evaluator")
    parser.add_argument("--output", type=Path, required=True, help="Destination .xlsx path")
    parser.add_argument("--summary-output", type=Path, help="Optional concise text summary destination")
    parser.add_argument("--expected-samples", type=int, default=None)
    parser.add_argument("--expected-tasks", type=int, default=None)
    return parser.parse_args()


def aggregate_task_scores(samples: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    for index, sample in enumerate(samples):
        if sample.get("error"):
            raise ValueError(f"sample {index} contains a scorer error: {sample['error']}")

        try:
            key = (str(sample["split"]), str(sample["category"]), str(sample["task_name"]))
            score = float(sample["score"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"sample {index} is missing valid split/category/task_name/score fields") from exc
        grouped[key].append(score)

    split_order = {"In_Domain": 0, "Out_of_Domain": 1}
    rows = []
    for (split, category, task_name), scores in grouped.items():
        rows.append(
            {
                "split": split,
                "category": category,
                "task_name": task_name,
                "sample_count": len(scores),
                "average_score": sum(scores) / len(scores),
            }
        )
    return sorted(rows, key=lambda row: (split_order.get(row["split"], 2), row["task_name"]))


def _format_sheet(sheet: Any, widths: list[int]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def export_workbook(result: dict[str, Any], output_path: Path, expected_tasks: int | None = None) -> int:
    samples = result.get("samples")
    if not isinstance(samples, list) or not samples:
        raise ValueError("result JSON contains no samples")

    task_rows = aggregate_task_scores(samples)
    if expected_tasks is not None and len(task_rows) != expected_tasks:
        raise ValueError(f"expected {expected_tasks} tasks, found {len(task_rows)}")

    workbook = Workbook()
    task_sheet = workbook.active
    task_sheet.title = "Task Scores"
    task_sheet.append(["Split", "Category", "Task Name", "Sample Count", "Average Score"])
    for row in task_rows:
        task_sheet.append(
            [
                row["split"].replace("_", "-"),
                row["category"],
                row["task_name"],
                row["sample_count"],
                row["average_score"],
            ]
        )
    for cell in task_sheet["E"][1:]:
        cell.number_format = "0.000000"
    _format_sheet(task_sheet, [18, 20, 78, 15, 16])

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Metric", "Sample Count", "Average Score"])
    summary = result.get("summary", {})
    for key, label in (("overall", "Overall"), ("In_Domain", "In-Domain"), ("Out_of_Domain", "Out-of-Domain")):
        item = summary.get(key, {})
        summary_sheet.append([label, item.get("num_samples"), item.get("mean_score")])
    for category, score in sorted(summary.get("overall", {}).get("by_category", {}).items()):
        summary_sheet.append([f"Category: {category}", None, score])
    for cell in summary_sheet["C"][1:]:
        cell.number_format = "0.000000"
    _format_sheet(summary_sheet, [32, 16, 16])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(task_rows)


def export_summary_text(result: dict[str, Any], output_path: Path, task_count: int) -> None:
    summary = result.get("summary", {})
    try:
        overall = summary["overall"]
        in_domain = summary["In_Domain"]
        out_of_domain = summary["Out_of_Domain"]
        lines = [
            f"Overall:        {float(overall['mean_score']):.6f}",
            f"In-Domain:      {float(in_domain['mean_score']):.6f}",
            f"Out-of-Domain:  {float(out_of_domain['mean_score']):.6f}",
            "",
            f"Samples: {int(overall['num_samples'])} "
            f"({int(in_domain['num_samples'])} in-domain + {int(out_of_domain['num_samples'])} out-of-domain)",
            f"Tasks:   {task_count}",
        ]
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError("result JSON contains an invalid domain summary") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n")


def main() -> None:
    args = parse_args()
    result = json.loads(args.result_json.read_text())
    samples = result.get("samples", [])
    if args.expected_samples is not None and len(samples) != args.expected_samples:
        raise ValueError(f"expected {args.expected_samples} samples, found {len(samples)}")
    task_count = export_workbook(result, args.output, expected_tasks=args.expected_tasks)
    print(f"Wrote {task_count} task averages from {len(samples)} samples to {args.output}")
    if args.summary_output is not None:
        export_summary_text(result, args.summary_output, task_count)
        print(f"Wrote concise score summary to {args.summary_output}")


if __name__ == "__main__":
    main()
