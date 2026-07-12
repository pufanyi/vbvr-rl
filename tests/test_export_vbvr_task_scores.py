from openpyxl import load_workbook
from pytest import approx

from src.cli.export_vbvr_task_scores import aggregate_task_scores, export_summary_text, export_workbook


def _result() -> dict:
    samples = [
        {
            "split": "In_Domain",
            "category": "Abstraction",
            "task_name": "G-1_first",
            "score": 0.2,
            "error": None,
        },
        {
            "split": "In_Domain",
            "category": "Abstraction",
            "task_name": "G-1_first",
            "score": 0.8,
            "error": None,
        },
        {
            "split": "Out_of_Domain",
            "category": "Knowledge",
            "task_name": "G-2_second",
            "score": 0.4,
            "error": None,
        },
    ]
    return {
        "samples": samples,
        "summary": {
            "overall": {
                "num_samples": 3,
                "mean_score": sum(sample["score"] for sample in samples) / 3,
                "by_category": {"Abstraction": 0.5, "Knowledge": 0.4},
            },
            "In_Domain": {"num_samples": 2, "mean_score": 0.5},
            "Out_of_Domain": {"num_samples": 1, "mean_score": 0.4},
        },
    }


def test_aggregate_task_scores() -> None:
    rows = aggregate_task_scores(_result()["samples"])

    assert len(rows) == 2
    assert rows[0] == {
        "split": "In_Domain",
        "category": "Abstraction",
        "task_name": "G-1_first",
        "sample_count": 2,
        "average_score": 0.5,
    }


def test_export_workbook(tmp_path) -> None:
    output_path = tmp_path / "scores.xlsx"

    task_count = export_workbook(_result(), output_path, expected_tasks=2)

    assert task_count == 2
    workbook = load_workbook(output_path, data_only=True)
    assert workbook.sheetnames == ["Task Scores", "Summary"]
    assert workbook["Task Scores"].max_row == 3
    assert workbook["Task Scores"]["E2"].value == 0.5
    assert workbook["Summary"]["C2"].value == approx(_result()["summary"]["overall"]["mean_score"])


def test_export_summary_text(tmp_path) -> None:
    output_path = tmp_path / "final_scores.txt"

    export_summary_text(_result(), output_path, task_count=2)

    assert output_path.read_text() == (
        "Overall:        0.466667\n"
        "In-Domain:      0.500000\n"
        "Out-of-Domain:  0.400000\n"
        "\n"
        "Samples: 3 (2 in-domain + 1 out-of-domain)\n"
        "Tasks:   2\n"
    )
