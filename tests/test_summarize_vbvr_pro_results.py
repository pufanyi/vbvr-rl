import json
from pathlib import Path

from openpyxl import load_workbook
from pytest import raises

from src.cli.summarize_vbvr_pro_results import _load_run, discover_runs, generate_reports
from src.eval.evaluation_provenance import build_manifest, write_manifest


def _write_result(
    root: Path,
    run_name: str,
    task_a: tuple[float, float],
    task_b: tuple[float, float],
    *,
    result_name: str = "prepared_1024x1024_max5s_vbvr_results.json",
    scorer_hash: str = "b" * 64,
) -> Path:
    samples = []
    for task_name, split, category, scores in (
        ("G-1_task", "In_Domain", "Abstraction", task_a),
        ("G-2_task", "Out_of_Domain", "Knowledge", task_b),
    ):
        for index, score in enumerate(scores):
            samples.append(
                {
                    "task_name": task_name,
                    "split": split,
                    "category": category,
                    "score": score,
                    "video_file": f"{index}.mp4",
                    "error": None,
                }
            )
    task_scores = {"G-1_task": sum(task_a) / 2, "G-2_task": sum(task_b) / 2}
    in_score = task_scores["G-1_task"]
    out_score = task_scores["G-2_task"]
    result = {
        "samples": samples,
        "summary": {
            "In_Domain": {"mean_score": in_score, "num_samples": 2},
            "Out_of_Domain": {"mean_score": out_score, "num_samples": 2},
            "overall": {
                "mean_score": (in_score + out_score) / 2,
                "num_samples": 4,
                "by_task": task_scores,
                "by_category": {"Abstraction": in_score, "Knowledge": out_score},
            },
        },
    }
    run_dir = root / run_name
    path = run_dir / "scores" / result_name
    path.parent.mkdir(parents=True)
    path.write_text(json.dumps(result), encoding="utf-8")
    provenance = build_manifest(
        stage="vbvr-pro-score",
        values={
            "state": "complete",
            "evalkit_revision": "a" * 40,
            "evalkit_source_sha256": scorer_hash,
        },
        files={},
        trees={},
        output_files={"result": str(path)},
    )
    write_manifest(run_dir / "score-provenance.json", provenance)
    return path


def test_load_run_rejects_result_replaced_after_provenance(tmp_path: Path) -> None:
    root = tmp_path / "results"
    result_path = _write_result(root, "unipc", (0.2, 0.4), (0.6, 0.8))
    result_path.write_text(result_path.read_text(encoding="utf-8") + "\n", encoding="utf-8")

    with raises(ValueError, match="recorded artifact changed"):
        _load_run(root / "unipc", expected_samples=4)


def test_generate_reports_accepts_dynamic_result_names_and_generic_cells(tmp_path: Path) -> None:
    root = tmp_path / "results"
    _write_result(root, "unipc", (0.4, 0.6), (0.5, 0.7), result_name="unipc_vbvr_results.json")
    _write_result(root, "euler", (0.2, 0.4), (0.6, 0.8), result_name="euler_vbvr_results.json")
    _write_result(root, "cps-noise-0.7", (0.6, 0.8), (0.3, 0.5))
    (root / "cps-noise-0.3" / "scores").mkdir(parents=True)
    output = tmp_path / "reports"

    workbook_path, json_path, text_path, complete_count, skipped_count = generate_reports(
        root, output, expected_samples=4
    )

    assert (complete_count, skipped_count) == (3, 1)
    workbook = load_workbook(workbook_path, data_only=True)
    assert workbook.sheetnames == ["Runs", "Task Scores", "Skipped"]
    run_rows = list(workbook["Runs"].iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in run_rows] == ["cps-noise-0.7", "euler", "unipc"]
    assert workbook["Task Scores"].max_row == 7
    assert workbook["Skipped"]["A2"].value == "cps-noise-0.3"

    summary = json.loads(json_path.read_text(encoding="utf-8"))
    assert summary["run_count"] == 3
    assert summary["skipped_count"] == 1
    assert summary["evalkit_source_sha256"] == "b" * 64
    assert [run["name"] for run in summary["runs"]] == ["cps-noise-0.7", "euler", "unipc"]
    assert text_path.read_text(encoding="utf-8").startswith("Run\tOverall\tIn-Domain")


def test_discover_runs_refuses_mixed_scorer_contracts(tmp_path: Path) -> None:
    root = tmp_path / "results"
    _write_result(root, "unipc", (0.2, 0.4), (0.6, 0.8), scorer_hash="b" * 64)
    _write_result(root, "euler", (0.2, 0.4), (0.6, 0.8), scorer_hash="c" * 64)

    with raises(ValueError, match="multiple scorer fingerprints"):
        discover_runs(root, expected_samples=4)


def test_discover_runs_accepts_one_run_as_root(tmp_path: Path) -> None:
    root = tmp_path / "results"
    _write_result(root, "unipc", (0.2, 0.4), (0.6, 0.8))

    runs, skipped = discover_runs(root / "unipc", expected_samples=4)

    assert [run.name for run in runs] == ["unipc"]
    assert skipped == []
